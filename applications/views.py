from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Q
from django.core.paginator import Paginator
from .models import Application, Document, DocumentType
from .forms import ApplicationForm, DocumentUploadForm, DocumentReviewForm, ApplicationStatusForm
from core.models import AdmissionCampaign, University, Faculty, Specialty


# ========== ABITURIYENT VIEWS ==========

@login_required
def create_application(request):
    active_campaign = AdmissionCampaign.objects.filter(faol=True).first()
    if not active_campaign:
        messages.warning(request, "Hozirda faol qabul davri mavjud emas.")
        return redirect('dashboard')

    if request.method == 'POST':
        form = ApplicationForm(request.POST)
        if form.is_valid():
            application = form.save(commit=False)
            application.abituriyent = request.user
            application.qabul_davri = active_campaign
            application.save()
            messages.success(request, "Ariza muvaffaqiyatli yaratildi! Endi hujjatlarni yuklang.")
            return redirect('upload_documents', pk=application.pk)
    else:
        form = ApplicationForm()

    faculties = Faculty.objects.filter(faol=True).prefetch_related('yonalishlar')
    return render(request, 'applications/create_application.html', {
        'form': form,
        'active_campaign': active_campaign,
        'faculties': faculties,
    })


@login_required
def application_detail(request, pk):
    application = get_object_or_404(Application, pk=pk, abituriyent=request.user)
    documents = application.hujjatlar.all()
    document_types = DocumentType.objects.all()
    uploaded_types = documents.values_list('hujjat_turi_id', flat=True)
    missing_types = document_types.filter(majburiy=True).exclude(id__in=uploaded_types)

    return render(request, 'applications/application_detail.html', {
        'application': application,
        'documents': documents,
        'missing_types': missing_types,
    })


@login_required
def upload_documents(request, pk):
    application = get_object_or_404(Application, pk=pk, abituriyent=request.user)

    if request.method == 'POST':
        form = DocumentUploadForm(request.POST, request.FILES)
        if form.is_valid():
            document = form.save(commit=False)
            document.ariza = application
            document.save()
            messages.success(request, f"'{document.hujjat_turi.nomi}' muvaffaqiyatli yuklandi!")
            return redirect('upload_documents', pk=pk)
    else:
        form = DocumentUploadForm()

    documents = application.hujjatlar.all()
    document_types = DocumentType.objects.all()
    uploaded_types = documents.values_list('hujjat_turi_id', flat=True)
    missing_types = document_types.filter(majburiy=True).exclude(id__in=uploaded_types)

    return render(request, 'applications/upload_documents.html', {
        'application': application,
        'form': form,
        'documents': documents,
        'missing_types': missing_types,
        'document_types': document_types,
    })


@login_required
def delete_document(request, pk):
    document = get_object_or_404(Document, pk=pk, ariza__abituriyent=request.user)
    app_pk = document.ariza.pk
    if document.holat in ['yuklangan', 'rad_etildi']:
        document.fayl.delete()
        document.delete()
        messages.success(request, "Hujjat o'chirildi.")
    else:
        messages.error(request, "Bu hujjatni o'chirish mumkin emas.")
    return redirect('upload_documents', pk=app_pk)


# ========== PUBLIC VIEWS ==========

def check_status(request):
    """Public page to check application status by ID."""
    application = None
    searched = False
    if request.GET.get('ariza_id'):
        searched = True
        try:
            ariza_id = int(request.GET.get('ariza_id'))
            application = Application.objects.select_related(
                'yonalish', 'yonalish__fakultet', 'qabul_davri'
            ).get(pk=ariza_id)
        except (ValueError, Application.DoesNotExist):
            application = None
    return render(request, 'applications/check_status.html', {
        'application': application,
        'searched': searched,
    })


# ========== PRINT VIEW ==========

@login_required
def print_application(request, pk):
    """Print-friendly application view."""
    if hasattr(request.user, 'staff_profile'):
        application = get_object_or_404(
            Application.objects.select_related(
                'abituriyent', 'abituriyent__applicant_profile',
                'yonalish', 'yonalish__fakultet', 'qabul_davri'
            ), pk=pk
        )
    else:
        application = get_object_or_404(
            Application.objects.select_related(
                'abituriyent', 'abituriyent__applicant_profile',
                'yonalish', 'yonalish__fakultet', 'qabul_davri'
            ), pk=pk, abituriyent=request.user
        )
    documents = application.hujjatlar.select_related('hujjat_turi').all()
    return render(request, 'applications/print_application.html', {
        'application': application,
        'documents': documents,
    })


# ========== XODIM/ADMIN VIEWS ==========

@login_required
def staff_dashboard(request):
    if not hasattr(request.user, 'staff_profile'):
        messages.error(request, "Sizda xodim huquqi yo'q.")
        return redirect('dashboard')

    total_applications = Application.objects.count()
    new_applications = Application.objects.filter(holat='yangi').count()
    reviewing_applications = Application.objects.filter(holat='korib_chiqilmoqda').count()
    accepted_applications = Application.objects.filter(holat='qabul_qilindi').count()
    rejected_applications = Application.objects.filter(holat='rad_etildi').count()

    # Faculty statistics
    faculty_stats = Faculty.objects.filter(faol=True).annotate(
        arizalar_soni=Count('yonalishlar__arizalar')
    )

    # Recent applications
    recent_applications = Application.objects.select_related(
        'abituriyent', 'yonalish', 'yonalish__fakultet'
    ).order_by('-yaratilgan_sana')[:10]

    # Status distribution for chart
    status_data = {
        'labels': ['Yangi', "Ko'rib chiqilmoqda", 'Qabul qilindi', 'Rad etildi'],
        'data': [new_applications, reviewing_applications, accepted_applications, rejected_applications],
        'colors': ['#3b82f6', '#f59e0b', '#10b981', '#ef4444'],
    }

    context = {
        'total_applications': total_applications,
        'new_applications': new_applications,
        'reviewing_applications': reviewing_applications,
        'accepted_applications': accepted_applications,
        'rejected_applications': rejected_applications,
        'faculty_stats': faculty_stats,
        'recent_applications': recent_applications,
        'status_data': status_data,
    }
    return render(request, 'applications/staff_dashboard.html', context)


@login_required
def staff_applications_list(request):
    if not hasattr(request.user, 'staff_profile'):
        return redirect('dashboard')

    applications = Application.objects.select_related(
        'abituriyent', 'yonalish', 'yonalish__fakultet', 'qabul_davri'
    )

    # Filters
    holat = request.GET.get('holat')
    fakultet_id = request.GET.get('fakultet')
    qidiruv = request.GET.get('q')

    if holat:
        applications = applications.filter(holat=holat)
    if fakultet_id:
        applications = applications.filter(yonalish__fakultet_id=fakultet_id)
    if qidiruv:
        applications = applications.filter(
            Q(abituriyent__first_name__icontains=qidiruv) |
            Q(abituriyent__last_name__icontains=qidiruv) |
            Q(abituriyent__applicant_profile__passport_seriya__icontains=qidiruv)
        )

    # Pagination
    paginator = Paginator(applications, 20)
    page_number = request.GET.get('sahifa')
    page_obj = paginator.get_page(page_number)

    faculties = Faculty.objects.filter(faol=True)

    return render(request, 'applications/staff_applications_list.html', {
        'page_obj': page_obj,
        'faculties': faculties,
        'current_holat': holat,
        'current_fakultet': fakultet_id,
        'qidiruv': qidiruv or '',
    })


@login_required
def staff_application_detail(request, pk):
    if not hasattr(request.user, 'staff_profile'):
        return redirect('dashboard')

    application = get_object_or_404(
        Application.objects.select_related(
            'abituriyent', 'abituriyent__applicant_profile',
            'yonalish', 'yonalish__fakultet'
        ),
        pk=pk
    )
    documents = application.hujjatlar.select_related('hujjat_turi').all()

    if request.method == 'POST':
        status_form = ApplicationStatusForm(request.POST)
        if status_form.is_valid():
            application.holat = status_form.cleaned_data['holat']
            if status_form.cleaned_data['izoh']:
                application.izoh = status_form.cleaned_data['izoh']
            application.save()
            messages.success(request, f"Ariza holati '{application.get_holat_display()}' ga o'zgartirildi.")
            return redirect('staff_application_detail', pk=pk)
    else:
        status_form = ApplicationStatusForm(initial={'holat': application.holat})

    return render(request, 'applications/staff_application_detail.html', {
        'application': application,
        'documents': documents,
        'status_form': status_form,
    })


@login_required
def staff_review_document(request, pk):
    if not hasattr(request.user, 'staff_profile'):
        return redirect('dashboard')

    document = get_object_or_404(Document.objects.select_related('ariza', 'hujjat_turi'), pk=pk)

    if request.method == 'POST':
        form = DocumentReviewForm(request.POST)
        if form.is_valid():
            document.holat = form.cleaned_data['holat']
            document.izoh = form.cleaned_data['izoh']
            document.tekshirgan = request.user
            document.save()
            messages.success(request, f"Hujjat {document.get_holat_display().lower()}.")
            return redirect('staff_application_detail', pk=document.ariza.pk)
    else:
        form = DocumentReviewForm()

    return render(request, 'applications/staff_review_document.html', {
        'document': document,
        'form': form,
    })


# ========== STAFF CRUD VIEWS ==========

@login_required
def staff_faculties(request):
    if not hasattr(request.user, 'staff_profile'):
        return redirect('dashboard')
    faculties = Faculty.objects.annotate(yonalishlar_soni=Count('yonalishlar'))
    return render(request, 'applications/staff_faculties.html', {'faculties': faculties})


@login_required
def staff_specialties(request):
    if not hasattr(request.user, 'staff_profile'):
        return redirect('dashboard')
    specialties = Specialty.objects.select_related('fakultet').all()
    faculties = Faculty.objects.filter(faol=True)

    fakultet_id = request.GET.get('fakultet')
    if fakultet_id:
        specialties = specialties.filter(fakultet_id=fakultet_id)

    return render(request, 'applications/staff_specialties.html', {
        'specialties': specialties,
        'faculties': faculties,
        'current_fakultet': fakultet_id,
    })


@login_required
def staff_document_types(request):
    if not hasattr(request.user, 'staff_profile'):
        return redirect('dashboard')
    document_types = DocumentType.objects.all()
    return render(request, 'applications/staff_document_types.html', {'document_types': document_types})


# ========== EXPORT ==========

@login_required
def export_applications_excel(request):
    """Export applications list to Excel."""
    if not hasattr(request.user, 'staff_profile'):
        return redirect('dashboard')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    applications = Application.objects.select_related(
        'abituriyent', 'abituriyent__applicant_profile',
        'yonalish', 'yonalish__fakultet', 'qabul_davri'
    )

    # Apply same filters as list view
    holat = request.GET.get('holat')
    fakultet_id = request.GET.get('fakultet')
    if holat:
        applications = applications.filter(holat=holat)
    if fakultet_id:
        applications = applications.filter(yonalish__fakultet_id=fakultet_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Arizalar"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    headers = [
        '#', 'F.I.O', 'Passport', 'Telefon',
        "Yo'nalish", 'Fakultet', "Ta'lim turi",
        'Holat', 'Hujjatlar soni', 'Sana',
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    status_labels = dict(Application.STATUS_CHOICES)
    for row, app in enumerate(applications, 2):
        profile = getattr(app.abituriyent, 'applicant_profile', None)
        cells_data = [
            app.pk,
            app.abituriyent.get_full_name(),
            getattr(profile, 'passport_seriya', '') if profile else '',
            getattr(profile, 'telefon', '') if profile else '',
            app.yonalish.nomi,
            app.yonalish.fakultet.nomi,
            app.yonalish.get_talim_turi_display(),
            status_labels.get(app.holat, app.holat),
            app.hujjatlar_soni,
            app.yaratilgan_sana.strftime('%d.%m.%Y'),
        ]
        for col, value in enumerate(cells_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    # Auto-adjust column widths
    for col_cells in ws.columns:
        max_length = 0
        for cell in col_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except TypeError:
                pass
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_length + 4, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="arizalar.xlsx"'
    wb.save(response)
    return response


# ========== API ==========

@login_required
def api_stats(request):
    if not hasattr(request.user, 'staff_profile'):
        return JsonResponse({'error': "Ruxsat yo'q"}, status=403)

    faculty_data = list(
        Faculty.objects.filter(faol=True).annotate(
            soni=Count('yonalishlar__arizalar')
        ).values('nomi', 'soni')
    )

    return JsonResponse({
        'faculty_data': faculty_data,
    })
